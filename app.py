import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from io import BytesIO
import matplotlib.pyplot as plt
import numpy as np

# Function to process each area based on the date column and starting row
def process_area(sheet, date_column, date_row, start_row, location):
    date = sheet[f'{date_column}{date_row}'].value
    am_names = [sheet[f'{date_column}{i}'].value for i in range(start_row, start_row + 10)]
    pm_names = [sheet[f'{date_column}{i}'].value for i in range(start_row + 10, start_row + 20)]
    names = [(name, 'AM') for name in am_names] + [(name, 'PM') for name in pm_names]

    processed_data = []
    for name, type_ in names:
        if name:
            preceptor, student = (name.split(' ~ ') if ' ~ ' in name else (name, None))
            student_placed = 'Yes' if student else 'No'
            student_type = None
            if student:
                if '(MD)' in student:
                    student_type = 'MD'
                elif '(PA)' in student:
                    student_type = 'PA'

            processed_data.append({
                'Date': date,
                'Type': type_,
                'Description': name,
                'Preceptor': preceptor.strip(),
                'Student': student.strip() if student else None,
                'Student Placed': student_placed,
                'Student Type': student_type,
                'Location': location
            })
    return processed_data

def correct_student_designations(df):
    """
    Corrects student designations (e.g., missing (MD) or (PA)) in the dataset
    by checking for matching names with proper designations elsewhere in the dataset.
    """
    # Extract students with and without designations
    df['Student Designation'] = df['Student'].str.extract(r'\((MD|PA)\)')
    students_with_designations = df.dropna(subset=['Student Designation'])
    students_without_designations = df[df['Student Designation'].isna()]

    # Map correct designations for students without them
    for _, row in students_without_designations.iterrows():
        student_name = row['Student'].strip() if row['Student'] else None

        if student_name:
            # Check if a matching name exists in rows with designations
            matches = students_with_designations[
                students_with_designations['Student'].str.contains(student_name, na=False, case=False, regex=False)
            ]
            if not matches.empty:
                # If a match is found, append the first correct designation
                correct_designation = matches['Student Designation'].iloc[0]
                df.loc[row.name, 'Student'] = f"{student_name} ({correct_designation})"

    # Drop the temporary column used for correction
    df.drop(columns=['Student Designation'], inplace=True)

    return df

# Streamlit app
st.title('OPD Data Processor')

uploaded_files = st.file_uploader("Choose Excel files", type="xlsx", accept_multiple_files=True)

if uploaded_files:
    all_data = []
    for uploaded_file in uploaded_files:
        wb = load_workbook(uploaded_file)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for date_column in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                for date_row in [4, 28, 52, 76]:
                    start_row = date_row + 2
                    area_data = process_area(sheet, date_column, date_row, start_row, sheet_name)
                    all_data.extend(area_data)

    df = pd.DataFrame(all_data)

    # Exclude rows with "COM CLOSED" or "Closed" in the Description column
    df = df[~df['Description'].str.contains('COM CLOSED|Closed', case=False, na=False)]

    # Add weekday column
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    df['Weekday'] = df['Date'].dt.day_name()

    # Calculate Days Worked by Preceptor
    filtered_df = df[df['Student Placed'] == 'Yes']
    filtered_df['Half Day'] = filtered_df['Type'].apply(lambda x: 0.5 if x in ['AM', 'PM'] else 0)
    days_worked = (
        filtered_df.groupby(['Preceptor', 'Date'])['Half Day']
        .sum()
        .reset_index()
        .rename(columns={'Half Day': 'Total Day Fraction'})
    )

    preceptor_days_summary = (
        days_worked.groupby('Preceptor')['Total Day Fraction']
        .sum()
        .reset_index()
        .rename(columns={'Total Day Fraction': 'Total Days'})
    )

    # Calculate available and used shifts
    available_shifts = (
        df.groupby(['Preceptor', 'Date', 'Type'])
        .size()
        .reset_index(name='Available Shifts')
    )
    available_shifts = (
        available_shifts.groupby('Preceptor')['Available Shifts']
        .sum()
        .reset_index()
    )

    used_shifts = (
        filtered_df.groupby(['Preceptor', 'Date', 'Type'])
        .size()
        .reset_index(name='Used Shifts')
    )
    used_shifts = (
        used_shifts.groupby('Preceptor')['Used Shifts']
        .sum()
        .reset_index()
    )

    # Merge shifts data for preceptor summary
    shifts_summary = pd.merge(available_shifts, used_shifts, on='Preceptor', how='left')
    shifts_summary['Used Shifts'] = shifts_summary['Used Shifts'].fillna(0)

    # Plot the graph for total days worked
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.bar(preceptor_days_summary['Preceptor'], preceptor_days_summary['Total Days'])
    ax.set_xlabel('Preceptor')
    ax.set_ylabel('Total Days Worked')
    ax.set_title('Total Days Worked by Preceptor')
    plt.xticks(rotation=45, fontsize=10, ha='right')
    st.pyplot(fig)

    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    df['Weekday'] = df['Date'].dt.day_name()

    # Sort weekdays (Monday to Sunday)
    weekday_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    df['Weekday'] = pd.Categorical(df['Weekday'], categories=weekday_order, ordered=True)

    # Calculate available and used shifts
    total_shifts = df.groupby(['Location', 'Weekday', 'Type']).size().reset_index(name='Total Shifts')
    open_shifts = df[df['Student Placed'] == 'No'].groupby(['Location', 'Weekday', 'Type']).size().reset_index(name='Open Shifts')

    # Merge total shifts and open shifts
    location_shifts = pd.merge(total_shifts, open_shifts, on=['Location', 'Weekday', 'Type'], how='left')
    location_shifts['Open Shifts'] = location_shifts['Open Shifts'].fillna(0)

    # Calculate percentage open shifts by location
    location_shifts['Percentage Open'] = (location_shifts['Open Shifts'] / location_shifts['Total Shifts']) * 100

    # Total percentage (all locations combined)
    total_shifts_summary = location_shifts.groupby(['Weekday', 'Type'])[['Total Shifts', 'Open Shifts']].sum().reset_index()
    total_shifts_summary['Percentage Open'] = (total_shifts_summary['Open Shifts'] / total_shifts_summary['Total Shifts']) * 100

    # Individual Preceptor Percentage
    preceptor_summary = df[df['Student Placed'] == 'Yes'].groupby(['Preceptor', 'Type'])['Student Placed'].count()
    preceptor_summary = preceptor_summary.div(
        df.groupby(['Preceptor', 'Type'])['Student Placed'].size()
    ).reset_index(name='Percentage Filled') * 100

    # Reshape data for display
    total_shifts_summary_pivot = total_shifts_summary.pivot(index='Weekday', columns='Type', values='Percentage Open').fillna(0).reset_index()

    # Plot the graph for available vs. used shifts
    fig, ax = plt.subplots(figsize=(12, 6))
    ax.bar(shifts_summary['Preceptor'], shifts_summary['Available Shifts'], label='Available Shifts', alpha=0.7)
    ax.bar(shifts_summary['Preceptor'], shifts_summary['Used Shifts'], label='Used Shifts', alpha=0.7)
    ax.set_xlabel('Preceptor')
    ax.set_ylabel('Shifts')
    ax.set_title('Available vs. Used Shifts by Preceptor')
    ax.legend()
    plt.xticks(rotation=45, fontsize=10, ha='right')
    st.pyplot(fig)

        # Plot total percentage open shifts by AM/PM
    fig, ax = plt.subplots(figsize=(10, 6))
    for shift_type in ['AM', 'PM']:
        ax.plot(total_shifts_summary_pivot['Weekday'], total_shifts_summary_pivot[shift_type], marker='o', label=shift_type)
    ax.set_title('Total Percentage of Open Shifts by Weekday (AM/PM)')
    ax.set_ylabel('Percentage Open Shifts')
    ax.set_xlabel('Weekday')
    plt.xticks(rotation=45)
    ax.legend()
    st.pyplot(fig)

    # Plot percentage open shifts by location and type
    locations = location_shifts['Location'].unique()
    fig, ax = plt.subplots(figsize=(12, 8))
    x = np.arange(len(weekday_order))
    bar_width = 0.2
    for i, location in enumerate(locations):
        data = location_shifts[location_shifts['Location'] == location]
        data = data.groupby(['Weekday', 'Type'])['Percentage Open'].mean().unstack()
        ax.bar(x + (i * bar_width), data['AM'], bar_width, label=f'{location} - AM')
        ax.bar(x + (i * bar_width) + (bar_width / 2), data['PM'], bar_width, label=f'{location} - PM')
    ax.set_title('Percentage of Open Shifts by Location and Type')
    ax.set_ylabel('Percentage Open Shifts')
    ax.set_xlabel('Weekday')
    ax.set_xticks(x + (len(locations) - 1) * bar_width / 2)
    ax.set_xticklabels(weekday_order)
    plt.xticks(rotation=45)
    ax.legend(loc='upper left', bbox_to_anchor=(1, 1))
    st.pyplot(fig)

    # Include all data in the downloadable Excel file
    output_file = BytesIO()
    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Combined Dataset')  # Full dataset
        preceptor_days_summary.to_excel(writer, index=False, sheet_name='Total Days Worked')  # Total days worked
        shifts_summary.to_excel(writer, index=False, sheet_name='Shifts Summary')  # Available vs. used shifts
    output_file.seek(0)

    st.download_button(
        label="Download Combined and Summary Data",
        data=output_file,
        file_name="combined_and_preceptor_data.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

